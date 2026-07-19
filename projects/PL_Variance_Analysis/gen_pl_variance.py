"""
NovaStar Financial Group — P&L Variance Data Generator (vectorized)
60 branches × 40 accounts × 365 days × 3 years ≈ 2,628,000 actual rows
"""
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, json

np.random.seed(42)
OUT  = "/sessions/zen-zealous-lovelace/mnt/outputs/PL_Variance_Analysis/data"
os.makedirs(OUT, exist_ok=True)

# ── DIMENSIONS ────────────────────────────────────────────────────────────────
accounts = pd.DataFrame([
    (1,"NII001","Net Interest Income - Loans",       "Revenue",  "Net Interest Income",  1, 3800),
    (2,"NII002","Net Interest Income - Mortgages",   "Revenue",  "Net Interest Income",  1, 2900),
    (3,"NII003","Net Interest Income - Deposits",    "Revenue",  "Net Interest Income",  1, 2100),
    (4,"NII004","Net Interest Income - Commercial",  "Revenue",  "Net Interest Income",  1, 2600),
    (5,"FEE001","Advisory & Management Fees",        "Revenue",  "Fee Income",           1, 1800),
    (6,"FEE002","Brokerage Commissions",             "Revenue",  "Fee Income",           1,  900),
    (7,"FEE003","Insurance Premiums Written",        "Revenue",  "Fee Income",           1, 1200),
    (8,"FEE004","Card Transaction Fees",             "Revenue",  "Fee Income",           1,  750),
    (9,"FEE005","Account Service Charges",           "Revenue",  "Fee Income",           1,  600),
   (10,"FEE006","Investment Banking Fees",           "Revenue",  "Fee Income",           1,  950),
   (11,"TRD001","Trading Revenue",                   "Revenue",  "Trading",              1, 1400),
   (12,"OTH001","Other Income",                      "Revenue",  "Other",                1,  450),
   (13,"PRV001","Loan Loss Provision - Consumer",    "Provision","Credit Losses",       -1,  420),
   (14,"PRV002","Loan Loss Provision - Commercial",  "Provision","Credit Losses",       -1,  310),
   (15,"PRV003","Insurance Claims Expense",          "Provision","Claims",              -1,  280),
   (16,"PER001","Compensation & Salaries",           "OpEx",     "Personnel",           -1, 2800),
   (17,"PER002","Performance Bonuses",               "OpEx",     "Personnel",           -1,  600),
   (18,"PER003","Benefits & Employee Costs",         "OpEx",     "Personnel",           -1,  550),
   (19,"TEC001","Core Banking Systems",              "OpEx",     "Technology",          -1,  480),
   (20,"TEC002","Digital & Mobile Platform",         "OpEx",     "Technology",          -1,  220),
   (21,"TEC003","Data & Analytics",                  "OpEx",     "Technology",          -1,  190),
   (22,"TEC004","Cybersecurity",                     "OpEx",     "Technology",          -1,  140),
   (23,"OPS001","Branch Operations",                 "OpEx",     "Operations",          -1,  620),
   (24,"OPS002","Transaction Processing",            "OpEx",     "Operations",          -1,  310),
   (25,"OPS003","Compliance & Regulatory",           "OpEx",     "Operations",          -1,  280),
   (26,"OPS004","Audit & Risk Management",           "OpEx",     "Operations",          -1,  180),
   (27,"FAC001","Rent & Occupancy",                  "OpEx",     "Premises",            -1,  520),
   (28,"FAC002","Facilities Management",             "OpEx",     "Premises",            -1,  160),
   (29,"MKT001","Brand & Marketing",                 "OpEx",     "Marketing",           -1,  230),
   (30,"MKT002","Customer Acquisition Cost",         "OpEx",     "Marketing",           -1,  310),
   (31,"MKT003","Product Marketing",                 "OpEx",     "Marketing",           -1,  120),
   (32,"PRF001","Legal & Compliance Consulting",     "OpEx",     "Professional Svcs",   -1,  180),
   (33,"PRF002","Accounting & External Audit",       "OpEx",     "Professional Svcs",   -1,  140),
   (34,"PRF003","Management Consulting",             "OpEx",     "Professional Svcs",   -1,  210),
   (35,"ADM001","Corporate G&A",                     "OpEx",     "Admin",               -1,  290),
   (36,"ADM002","Business Insurance",                "OpEx",     "Admin",               -1,  120),
   (37,"ADM003","Communications & Utilities",        "OpEx",     "Admin",               -1,  110),
   (38,"DDA001","Depreciation",                      "D&A",      "D&A",                 -1,  380),
   (39,"DDA002","Amortization of Intangibles",       "D&A",      "D&A",                 -1,  140),
   (40,"TAX001","Corporate Income Tax",              "Tax",       "Tax",                 -1,  880),
], columns=["account_id","account_code","account_name","category","subcategory","p_l_sign","base_daily"])

# 60 branches: 12 per region × 5 regions
regions_list = ["Northeast","Southeast","Midwest","Southwest","West"]
cities = {
    "Northeast": ["New York","Boston","Philadelphia","Hartford","Providence","Albany"],
    "Southeast": ["Atlanta","Miami","Charlotte","Nashville","Tampa","Raleigh"],
    "Midwest":   ["Chicago","Detroit","Minneapolis","Columbus","Indianapolis","St Louis"],
    "Southwest": ["Dallas","Houston","Phoenix","San Antonio","Austin","El Paso"],
    "West":      ["Los Angeles","San Francisco","Seattle","Denver","Portland","Las Vegas"],
}
branch_rows = []
for rid, region in enumerate(regions_list):
    for i in range(12):
        city = cities[region][i % 6]
        suf  = "" if i < 6 else f" ({i-5})"
        branch_rows.append({
            "branch_id":   f"B{rid*12+i+1:03d}",
            "branch_name": f"{city} Branch{suf}",
            "region":      region,
            "city":        city,
            "branch_type": ["Corporate Hub","Regional Office","Standard Branch","Digital Branch"][min(i,3)],
            "perf_factor": round(np.random.uniform(0.82, 1.18), 3),
            "opened_year": int(np.random.choice([2008,2010,2012,2014,2016,2018,2020])),
        })
branches = pd.DataFrame(branch_rows)

print(f"Accounts: {len(accounts)} | Branches: {len(branches)}")
total_actual = len(branches) * len(accounts) * 365 * 3
print(f"Expected actual rows: {total_actual:,}")

# ── VECTORIZED ACTUAL GENERATOR ───────────────────────────────────────────────
# Seasonal factors: (12 months) × (account)
# Using category-level seasonal patterns
def make_seasonal(cat, subcat):
    if cat == "Revenue":
        if "Trading" in subcat:
            return [1.15,1.10,1.05,0.95,0.92,0.90,0.92,0.95,1.00,1.08,1.05,1.00]
        elif "Fee" in subcat:
            return [1.08,1.03,1.05,0.98,0.95,0.92,0.90,0.95,1.00,1.05,1.08,1.12]
        else:   # NII
            return [0.97,0.97,0.98,0.99,1.00,1.01,1.02,1.03,1.02,1.01,1.00,0.99]
    elif cat == "Provision":
        return [1.10,1.05,1.00,0.95,0.95,0.98,1.05,1.05,1.00,0.95,0.95,1.00]
    elif cat == "OpEx":
        if "Personnel" in subcat:
            return [1.25,0.95,0.95,0.95,0.95,0.95,0.95,0.95,0.95,0.95,0.95,1.25]
        elif "Marketing" in subcat:
            return [1.05,1.00,1.05,1.08,1.10,1.05,0.95,0.90,0.95,1.00,1.05,1.10]
        else:
            return [1.02,0.98,0.99,1.00,1.00,1.01,1.01,1.00,1.00,1.00,1.00,1.01]
    else:
        return [1.0]*12

# Build seasonal matrix: shape (40 accounts, 12 months)
seas_matrix = np.array([
    make_seasonal(r.category, r.subcategory) for _, r in accounts.iterrows()
])  # (40, 12)

# Year-over-year growth by category
def yoy(year, cat):
    if year == 2022: return 1.0
    growth = {"Revenue":1.07,"Provision":1.15,"OpEx":1.10,"D&A":1.08,"Tax":1.06}
    g2023 = growth.get(cat, 1.07)
    g2024 = {"Revenue":1.09,"Provision":1.08,"OpEx":1.12,"D&A":1.10,"Tax":1.08}.get(cat,1.09)
    return g2023 if year == 2023 else g2023 * g2024

# Actual variance by category and year
def act_var_params(year, cat):
    """Returns (mean, std) for actual vs budget noise."""
    if year == 2022: return (0.0, 0.04)
    if year == 2023:
        return {"Revenue":(0.02,0.03),"OpEx":(-0.04,0.03),"Provision":(-0.02,0.03)}.get(cat,(0.0,0.03))
    # 2024: revenue beat, costs over
    return {"Revenue":(0.05,0.03),"OpEx":(-0.08,0.04),"Provision":(-0.06,0.03)}.get(cat,(-0.03,0.02))

all_actual_dfs = []

for year in [2022, 2023, 2024]:
    dates = pd.date_range(f"{year}-01-01", f"{year}-12-31", freq="D")
    n_days = len(dates)
    n_branches = len(branches)
    n_accounts = len(accounts)
    N = n_branches * n_accounts * n_days

    print(f"\n▶ {year}: {n_branches}br × {n_accounts}acct × {n_days}d = {N:,} rows", flush=True)

    # Create index arrays
    b_idx = np.repeat(np.arange(n_branches), n_accounts * n_days)   # branch index
    a_idx = np.tile(np.repeat(np.arange(n_accounts), n_days), n_branches)  # account index
    d_idx = np.tile(np.arange(n_days), n_branches * n_accounts)     # day index

    # Base daily amounts (signed) for each account
    base_signed = accounts["base_daily"].values * accounts["p_l_sign"].values  # (40,)

    # Seasonal factor per account per day
    months = dates.month.values  # (n_days,)
    seas_by_day = seas_matrix[:, months - 1]  # (40, n_days)

    # YoY growth per account
    yoy_factors = np.array([yoy(year, cat) for cat in accounts["category"].values])  # (40,)

    # Weekday factor per day
    wd_map = {0:1.0, 1:1.0, 2:1.0, 3:1.0, 4:1.0, 5:0.60, 6:0.30}
    wd_factors = np.array([wd_map[d.dayofweek] for d in dates])  # (n_days,)

    # Branch perf factors
    br_perf = branches["perf_factor"].values  # (60,)

    # Build amount matrix: (n_branches, n_accounts, n_days)
    base_3d   = base_signed[None, :, None]            # (1, 40, 1)
    yoy_3d    = yoy_factors[None, :, None]             # (1, 40, 1)
    seas_3d   = seas_by_day[None, :, :]               # (1, 40, n_days)
    wd_3d     = wd_factors[None, None, :]             # (1, 1, n_days)
    perf_3d   = br_perf[:, None, None]                # (60, 1, 1)

    # Variance noise per account category and year
    cat_list = accounts["category"].values
    var_means = np.array([act_var_params(year, cat)[0] for cat in cat_list])  # (40,)
    var_stds  = np.array([act_var_params(year, cat)[1] for cat in cat_list])  # (40,)
    noise_3d  = np.random.normal(var_means[None, :, None], var_stds[None, :, None],
                                 size=(n_branches, n_accounts, n_days))
    daily_noise = 1 + noise_3d + np.random.normal(0, 0.015, size=(n_branches, n_accounts, n_days))

    amounts_3d = base_3d * yoy_3d * seas_3d * wd_3d * perf_3d * daily_noise

    # Flatten
    amounts_flat = amounts_3d.reshape(-1)

    df_year = pd.DataFrame({
        "date":       dates[d_idx].strftime("%Y-%m-%d"),
        "branch_id":  branches["branch_id"].values[b_idx],
        "account_id": accounts["account_id"].values[a_idx],
        "amount":     np.round(amounts_flat, 2),
    })

    out_path = f"{OUT}/actual_{year}.csv"
    df_year.to_csv(out_path, index=False)
    mb = os.path.getsize(out_path) / 1024 / 1024
    print(f"   Saved  actual_{year}.csv  |  {len(df_year):,} rows  |  {mb:.1f} MB", flush=True)
    all_actual_dfs.append(df_year)

# ── BUDGET DATA ────────────────────────────────────────────────────────────────
print("\n▶ Generating monthly budget...")
budget_rows = []
for year in [2022, 2023, 2024]:
    for month in range(1, 13):
        for _, br in branches.iterrows():
            for _, acct in accounts.iterrows():
                aid     = acct["account_id"]
                cat     = acct["category"]
                yg      = yoy(year, cat)
                sf      = make_seasonal(cat, acct["subcategory"])[month - 1]
                monthly = acct["base_daily"] * acct["p_l_sign"] * yg * sf * 22
                budget_rows.append((f"{year}-{month:02d}", year, month,
                                    br["branch_id"], aid, round(monthly, 2)))
df_budget = pd.DataFrame(budget_rows,
    columns=["period","year","month","branch_id","account_id","budget_amount"])
df_budget.to_csv(f"{OUT}/budget_monthly.csv", index=False)
print(f"   Budget: {len(df_budget):,} rows")

# ── EXCEL DIMENSIONS ──────────────────────────────────────────────────────────
print("\n▶ Building PL_Dimensions.xlsx...")
wb = Workbook()
DARK = "1E3A5F"; WHITE = "FFFFFF"; LGRAY = "F3F4F6"
hf  = Font(name="Calibri", bold=True, color=WHITE, size=10)
hfl = PatternFill("solid", fgColor=DARK)
bf  = Font(name="Calibri", size=9)
af  = PatternFill("solid", fgColor=LGRAY)
ca  = Alignment(horizontal="center", vertical="center")
la  = Alignment(horizontal="left",   vertical="center")
ts  = Side(style="thin", color="CCCCCC")
bdr = Border(left=ts, right=ts, top=ts, bottom=ts)

def write_ws(ws, df, title):
    n = len(df.columns)
    ws.merge_cells(f"A1:{get_column_letter(n)}1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", bold=True, size=13, color=DARK)
    ws["A1"].fill = PatternFill("solid", fgColor="DBEAFE")
    ws["A1"].alignment = ca
    ws.row_dimensions[1].height = 24
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=2, column=ci)
        c.value = col; c.font = hf; c.fill = hfl; c.alignment = ca; c.border = bdr
    ws.row_dimensions[2].height = 16
    for ri, row_vals in enumerate(df.itertuples(index=False), 3):
        fill = af if ri % 2 == 0 else PatternFill("solid", fgColor=WHITE)
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=ci)
            c.value = val; c.font = bf; c.fill = fill; c.border = bdr
            c.alignment = ca if isinstance(val, (int,float)) else la
        ws.row_dimensions[ri].height = 14

ws1 = wb.active; ws1.title = "dim_account"
write_ws(ws1, accounts.drop(columns=["base_daily"]),
         "NovaStar Financial — GL Account Dimension (40 accounts)")
for ci, w in enumerate([8,10,32,12,22,8], 1):
    ws1.column_dimensions[get_column_letter(ci)].width = w
ws1.freeze_panes = "A3"

ws2 = wb.create_sheet("dim_branch")
write_ws(ws2, branches.drop(columns=["perf_factor"]),
         "NovaStar Financial — Branch Dimension (60 branches)")
for ci, w in enumerate([8,24,14,16,18,12], 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.freeze_panes = "A3"

wb.save(f"{OUT}/PL_Dimensions.xlsx")
print(f"   Saved PL_Dimensions.xlsx")

# ── PRE-AGGREGATED STATS FOR HTML DASHBOARD ────────────────────────────────────
print("\n▶ Computing dashboard aggregations...")
df_all  = pd.concat(all_actual_dfs, ignore_index=True)
df_all["date"]   = pd.to_datetime(df_all["date"])
df_all["year"]   = df_all["date"].dt.year
df_all["month"]  = df_all["date"].dt.month
df_all["period"] = df_all["date"].dt.to_period("M").astype(str)
df_all = df_all.merge(accounts[["account_id","category","subcategory"]], on="account_id")
df_all = df_all.merge(branches[["branch_id","region"]], on="branch_id")

# Monthly P&L by category (actual)
mpl = df_all.groupby(["period","year","month","category"])["amount"].sum().reset_index()
mpl["amount"] = mpl["amount"].round(0).astype(int)

# Monthly budget by category
bcat = df_budget.merge(accounts[["account_id","category"]], on="account_id")
mbgt = bcat.groupby(["period","year","month","category"])["budget_amount"].sum().reset_index()
mbgt["budget_amount"] = mbgt["budget_amount"].round(0).astype(int)

cats = ["Revenue","Provision","OpEx","D&A","Tax"]
periods = sorted(mpl["period"].unique())
monthly_list = []
for p in periods:
    pa = mpl[mpl.period == p]
    pb = mbgt[mbgt.period == p]
    row = {"p": p}
    for c in cats:
        row[f"a_{c}"] = int(pa[pa.category == c]["amount"].sum())
        row[f"b_{c}"] = int(pb[pb.category == c]["budget_amount"].sum())
    monthly_list.append(row)

# Regional P&L 2024
r2024 = df_all[df_all.year == 2024].groupby(["region","category"])["amount"].sum().reset_index()
r2024["amount"] = r2024["amount"].round(0).astype(int)
reg_list = []
for region in regions_list:
    rd = r2024[r2024.region == region]
    row = {"region": region}
    for c in cats:
        row[f"a_{c}"] = int(rd[rd.category == c]["amount"].sum())
    reg_list.append(row)

# OpEx subcategory breakdown 2024
ox = df_all[(df_all.year==2024)&(df_all.category=="OpEx")].groupby("subcategory")["amount"].sum().reset_index()
ox["amount"] = ox["amount"].round(0).astype(int)
ox = ox.sort_values("amount").to_dict("records")

# Year summary
yr_summary = {}
for yr in [2022, 2023, 2024]:
    df_yr = df_all[df_all.year == yr]
    rev  = int(df_yr[df_yr.category=="Revenue"]["amount"].sum())
    prov = int(df_yr[df_yr.category=="Provision"]["amount"].sum())
    opex = int(df_yr[df_yr.category=="OpEx"]["amount"].sum())
    da   = int(df_yr[df_yr.category=="D&A"]["amount"].sum())
    tax  = int(df_yr[df_yr.category=="Tax"]["amount"].sum())
    ni   = rev + prov + opex + da + tax
    ebitda = rev + prov + opex
    yr_summary[yr] = {"rev":rev,"prov":prov,"opex":opex,"da":da,"tax":tax,"ni":ni,"ebitda":ebitda}
    print(f"  {yr}: Rev=${rev/1e6:.1f}M  OpEx=${opex/1e6:.1f}M  EBITDA=${ebitda/1e6:.1f}M  NI=${ni/1e6:.1f}M ({ni/rev*100:.1f}%)")

# 2024 budget year summary
b24 = df_budget[df_budget.year==2024].merge(accounts[["account_id","category"]], on="account_id")
b24g = b24.groupby("category")["budget_amount"].sum()
bgt_2024 = {c: int(b24g.get(c, 0)) for c in cats}
bgt_2024["ni"] = sum(bgt_2024.values())
bgt_2024["ebitda"] = bgt_2024["Revenue"] + bgt_2024["Provision"] + bgt_2024["OpEx"]

# Save aggregated data JSON for HTML dashboard
dashboard_data = {
    "monthly": monthly_list,
    "regional": reg_list,
    "opex_breakdown": ox,
    "year_summary": {str(k): v for k, v in yr_summary.items()},
    "budget_2024": bgt_2024,
}
with open(f"{OUT}/../dashboard_data.json", "w") as f:
    json.dump(dashboard_data, f, indent=2)

print("\n=== FINAL SUMMARY ===")
total = sum(876000 for _ in range(3))
actual_total = len(df_all)
print(f"Actual rows: {actual_total:,}")
print(f"Budget rows: {len(df_budget):,}")
print(f"Grand total: {actual_total + len(df_budget):,}")
ls_out = os.popen(f"ls -lh {OUT}/").read()
print(f"\nFiles:\n{ls_out}")
